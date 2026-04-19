"""
文部科学省「日本食品標準成分表2020年版（八訂）」Excel → JSON 変換スクリプト

使い方:
  1. https://www.mext.go.jp/a_menu/syokuhinseibun/mext_01110.html から
     「日本食品標準成分表2020年版（八訂）」本表 (xlsx) をダウンロード
  2. scripts/data/ ディレクトリに配置 (ファイル名例: 20230428-mxt_kagsei-mext_00001_012.xlsx)
  3. `python scripts/convert_mext.py <xlsx_path>` を実行
  4. data/mext_foods.json と data/mext_foods.search.json が出力される

依存:
  pip install openpyxl pykakasi

ライセンス:
  文部科学省の政府著作物 (出典明記で自由利用)
  出典: 文部科学省「日本食品標準成分表2020年版（八訂）」
"""
from __future__ import annotations

import argparse
import json
import re
import sys
from pathlib import Path
from typing import Any


# 成分表の列マッピング (2020年版八訂 本表 "表全体" シート、2021-12-28 更新版)
# inspect_mext.py で row 11 (成分識別子) を直接確認した列番号。
COLUMN_MAP: dict[str, int] = {
    "food_group": 0,      # 食品群番号
    "food_number": 1,     # 食品番号 ("01001")
    "name": 3,            # 食品名
    "kcal": 6,            # ENERC_KCAL
    "protein_g": 9,       # PROT-
    "fat_g": 12,          # FAT-
    "carb_g": 20,         # CHOCDF- (伝統的炭水化物、食物繊維含む)
    "fiber_g": 18,        # FIB-
    "potassium_mg": 24,   # K
    "calcium_mg": 25,     # CA
    "magnesium_mg": 26,   # MG
    "iron_mg": 28,        # FE
    "zinc_mg": 29,        # ZN
    "vitamin_a_ug": 42,   # VITA_RAE (レチノール活性当量)
    "vitamin_d_ug": 43,   # VITD
    "vitamin_e_mg": 44,   # TOCPHA (α-トコフェロール)
    "vitamin_k_ug": 48,   # VITK
    "vitamin_b1_mg": 49,  # THIA
    "vitamin_b2_mg": 50,  # RIBF
    "niacin_mg": 52,      # NE (ナイアシン当量)
    "vitamin_b6_mg": 53,  # VITB6A
    "vitamin_b12_ug": 54, # VITB12
    "folate_ug": 55,      # FOL
    "vitamin_c_mg": 58,   # VITC
    "salt_g": 60,         # 食塩相当量
}

DATA_START_ROW = 12  # 0-indexed。1行目(0)〜11行目(10)がヘッダ、12行目からデータ。


def parse_nutrient(value: Any) -> float | None:
    """MEXT表記ゆれを数値化。
    '-' / '' / None → None (未測定)
    'Tr' / '(Tr)'   → 0.0 (微量、0と近似)
    '(数値)'         → 数値 (推定値もそのまま採用)
    """
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip()
    if s in ("", "-"):
        return None
    if s in ("Tr", "(Tr)"):
        return 0.0
    m = re.match(r"^\(?(-?\d+(?:\.\d+)?)\)?$", s)
    if m:
        return float(m.group(1))
    return None


def to_hiragana(text: str) -> str:
    """カタカナ→ひらがな変換 (検索ヒット率向上)."""
    try:
        import pykakasi
    except ImportError:
        # pykakasi 未導入でも動くように素朴変換
        return "".join(
            chr(ord(c) - 0x60) if "ァ" <= c <= "ヶ" else c for c in text
        )
    kks = pykakasi.kakasi()
    result = kks.convert(text)
    return "".join(item["hira"] for item in result)


def normalize_name_for_search(name: str) -> str:
    """食品名から括弧内・記号を除いた検索用正規化."""
    s = re.sub(r"[()\[\]「」『』【】]", " ", name)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def convert(xlsx_path: Path, out_dir: Path) -> None:
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("ERROR: openpyxl が必要です。`pip install openpyxl pykakasi` を実行してください。", file=sys.stderr)
        sys.exit(1)

    print(f"Loading {xlsx_path}...")
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        print("ERROR: 最初のシートが読み取れませんでした。", file=sys.stderr)
        sys.exit(1)

    # 食品群番号 → 日本語名
    group_names = {
        "01": "穀類", "02": "いも及びでん粉類", "03": "砂糖及び甘味類",
        "04": "豆類", "05": "種実類", "06": "野菜類", "07": "果実類",
        "08": "きのこ類", "09": "藻類", "10": "魚介類", "11": "肉類",
        "12": "卵類", "13": "乳類", "14": "油脂類", "15": "菓子類",
        "16": "し好飲料類", "17": "調味料及び香辛料類", "18": "調理済み流通食品類",
    }

    foods: list[dict[str, Any]] = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < DATA_START_ROW:
            continue
        food_number = row[COLUMN_MAP["food_number"]] if COLUMN_MAP["food_number"] < len(row) else None
        name = row[COLUMN_MAP["name"]] if COLUMN_MAP["name"] < len(row) else None
        if food_number is None or name is None:
            continue
        food_id = str(food_number).strip().zfill(5)
        name_str = str(name).strip()
        if not name_str or not food_id.isdigit():
            continue

        nutrients: dict[str, float] = {}
        for key in (
            "kcal", "protein_g", "fat_g", "carb_g", "fiber_g", "salt_g",
            "calcium_mg", "iron_mg", "potassium_mg", "magnesium_mg", "zinc_mg",
            "vitamin_a_ug", "vitamin_d_ug", "vitamin_e_mg", "vitamin_k_ug",
            "vitamin_b1_mg", "vitamin_b2_mg", "niacin_mg", "vitamin_b6_mg",
            "vitamin_b12_ug", "folate_ug", "vitamin_c_mg",
        ):
            idx = COLUMN_MAP.get(key)
            if idx is None or idx >= len(row):
                continue
            v = parse_nutrient(row[idx])
            if v is not None:
                nutrients[key] = v

        if "kcal" not in nutrients:
            continue

        group_raw = row[COLUMN_MAP["food_group"]] if COLUMN_MAP["food_group"] < len(row) else None
        group_code = str(group_raw).strip().zfill(2) if group_raw else ""
        category = group_names.get(group_code, f"群{group_code}" if group_code else None)

        foods.append({
            "id": food_id,
            "name": name_str,
            "nameKana": to_hiragana(name_str),
            "category": category,
            "source": "mext",
            "nutrients": {
                "kcal": nutrients.get("kcal", 0),
                "protein_g": nutrients.get("protein_g", 0),
                "fat_g": nutrients.get("fat_g", 0),
                "carb_g": nutrients.get("carb_g", 0),
                **{k: v for k, v in nutrients.items() if k not in ("kcal", "protein_g", "fat_g", "carb_g")},
            },
        })

    print(f"Parsed {len(foods)} foods.")

    out_dir.mkdir(parents=True, exist_ok=True)
    full_path = out_dir / "mext_foods.json"
    search_path = out_dir / "mext_foods.search.json"

    with full_path.open("w", encoding="utf-8") as f:
        json.dump(foods, f, ensure_ascii=False, separators=(",", ":"))
    print(f"Wrote {full_path} ({full_path.stat().st_size // 1024} KB)")

    search_compact = [
        {
            "id": f["id"],
            "name": f["name"],
            "nameKana": f.get("nameKana"),
            "category": f.get("category"),
            "kcal": f["nutrients"].get("kcal", 0),
            "p": f["nutrients"].get("protein_g", 0),
            "f_": f["nutrients"].get("fat_g", 0),
            "c": f["nutrients"].get("carb_g", 0),
        }
        for f in foods
    ]
    with search_path.open("w", encoding="utf-8") as f:
        json.dump(search_compact, f, ensure_ascii=False, separators=(",", ":"))
    print(f"Wrote {search_path} ({search_path.stat().st_size // 1024} KB)")


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("xlsx", type=Path, help="MEXT 成分表 xlsx ファイルパス")
    parser.add_argument(
        "--out",
        type=Path,
        default=Path(__file__).resolve().parent.parent / "public" / "data",
        help="出力先ディレクトリ (デフォルト: public/data)",
    )
    args = parser.parse_args()
    if not args.xlsx.exists():
        print(f"ERROR: ファイルが見つかりません: {args.xlsx}", file=sys.stderr)
        sys.exit(1)
    convert(args.xlsx, args.out)


if __name__ == "__main__":
    main()
