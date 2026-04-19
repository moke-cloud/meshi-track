"""MEXT成分表 xlsx のシート構造・ヘッダ・最初の数行を調査するデバッグスクリプト."""
import sys
from pathlib import Path

from openpyxl import load_workbook


def main() -> None:
    path = Path(sys.argv[1])
    out_path = Path(sys.argv[2]) if len(sys.argv) > 2 else path.with_suffix(".inspect.txt")
    wb = load_workbook(path, read_only=True, data_only=True)
    lines: list[str] = []
    lines.append(f"Sheets: {wb.sheetnames}")
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        lines.append(f"\n=== Sheet: {sheet_name} | dims: {ws.max_row}rows x {ws.max_column}cols ===")
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i > 15:
                break
            truncated = [
                (str(v)[:30] + "...") if v is not None and len(str(v)) > 30 else v
                for v in row
            ]
            for col_idx, cell in enumerate(truncated):
                if cell is not None and str(cell).strip():
                    lines.append(f"  row {i:2d} col {col_idx:2d}: {cell!r}")
            lines.append("")
        break  # 最初のシートのみ (全表)
    out_path.write_text("\n".join(lines), encoding="utf-8")
    print(f"Wrote {out_path}")


if __name__ == "__main__":
    main()
